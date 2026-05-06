import io
import os
import re
import sqlite3
from contextlib import closing
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATABASE_URL = (
    os.environ.get("POSTGRES_URL")
    or os.environ.get("DATABASE_URL")
    or os.environ.get("POSTGRES_PRISMA_URL")
)
USE_PG = bool(DATABASE_URL)

if USE_PG:
    import psycopg
    PH = "%s"
    NOW_FN = "NOW()"
    PK_TYPE = "SERIAL PRIMARY KEY"
    TS_TYPE = "TIMESTAMPTZ"
    DATE_SELECT = "to_char(created_at, 'YYYY-MM-DD HH24:MI:SS')"
else:
    DB_PATH = BASE_DIR / "register.db"
    PH = "?"
    NOW_FN = "(datetime('now', 'localtime'))"
    PK_TYPE = "INTEGER PRIMARY KEY AUTOINCREMENT"
    TS_TYPE = "TEXT"
    DATE_SELECT = "created_at"

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
PHONE_RE = re.compile(r"^\d{8}$")
MAX_NAME = 64
MAX_EMAIL = 128

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def xlsx_safe(value):
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def db_connect():
    if USE_PG:
        return psycopg.connect(DATABASE_URL)
    return sqlite3.connect(DB_PATH)


def init_db() -> None:
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS register (
                id          {PK_TYPE},
                last_name   TEXT NOT NULL,
                first_name  TEXT NOT NULL,
                phone       TEXT NOT NULL,
                email       TEXT NOT NULL,
                created_at  {TS_TYPE} NOT NULL DEFAULT {NOW_FN}
            )
            """
        )
        conn.commit()


app = Flask(__name__)

try:
    init_db()
except Exception as e:
    print(f"init_db on startup failed (will retry on demand): {e}")


def fetch_rows():
    sql = (
        f"SELECT id, last_name, first_name, phone, email, {DATE_SELECT} "
        f"FROM register ORDER BY id DESC"
    )
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(sql)
        return [
            {
                "id": r[0],
                "last_name": r[1],
                "first_name": r[2],
                "phone": r[3],
                "email": r[4],
                "created_at": str(r[5]) if r[5] is not None else "",
            }
            for r in cur.fetchall()
        ]


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/admin")
def admin():
    return render_template("admin.html", rows=fetch_rows())


@app.post("/api/register")
def api_register():
    data = request.get_json(silent=True) or request.form
    last_name = (data.get("last_name") or "").strip()
    first_name = (data.get("first_name") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()

    if not last_name:
        return jsonify(ok=False, field="last_name", message="Овог оруулна уу."), 400
    if len(last_name) > MAX_NAME:
        return jsonify(ok=False, field="last_name", message=f"Овог {MAX_NAME} тэмдэгтээс хэтэрч болохгүй."), 400
    if not first_name:
        return jsonify(ok=False, field="first_name", message="Нэр оруулна уу."), 400
    if len(first_name) > MAX_NAME:
        return jsonify(ok=False, field="first_name", message=f"Нэр {MAX_NAME} тэмдэгтээс хэтэрч болохгүй."), 400
    if not PHONE_RE.match(phone):
        return jsonify(ok=False, field="phone", message="Утасны дугаар 8 оронтой тоо байх ёстой."), 400
    if not EMAIL_RE.match(email) or len(email) > MAX_EMAIL:
        return jsonify(
            ok=False,
            field="email",
            message="Имэйл буруу байна. Жишээ: example@gmail.com",
        ), 400

    sql = f"INSERT INTO register (last_name, first_name, phone, email) VALUES ({PH}, {PH}, {PH}, {PH})"
    if USE_PG:
        sql += " RETURNING id"
    with closing(db_connect()) as conn:
        cur = conn.cursor()
        cur.execute(sql, (last_name, first_name, phone, email))
        new_id = cur.fetchone()[0] if USE_PG else cur.lastrowid
        conn.commit()
    return jsonify(ok=True, id=new_id)


@app.get("/admin/export.xlsx")
def admin_export():
    rows = fetch_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Бүртгэл"

    headers = ["№", "Овог", "Нэр", "Утасны дугаар", "Имэйл", "Огноо"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F2A44")
    header_align = Alignment(vertical="center", horizontal="left")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([
            r["id"],
            xlsx_safe(r["last_name"]),
            xlsx_safe(r["first_name"]),
            xlsx_safe(r["phone"]),
            xlsx_safe(r["email"]),
            xlsx_safe(r["created_at"]),
        ])

    widths = [6, 22, 22, 16, 32, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y-%m-%d")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"register-{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
